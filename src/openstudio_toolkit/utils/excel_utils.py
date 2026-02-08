import logging

import numpy as np
import openpyxl
import pandas as pd
from openpyxl import Workbook, load_workbook

# Configure logger
logger = logging.getLogger(__name__)

def read_excel_table_to_df(
    excel_file_path: str, 
    sheet_name: str, 
    column_range: str | None = None, 
    start_row: int = 1, 
    end_row: int = 25000
) -> pd.DataFrame:
    """
    Read a specific table/range from an Excel sheet into a pandas DataFrame.

    Parameters:
    - excel_file_path (str): Absolute path to the Excel file.
    - sheet_name (str): Name of the sheet to read from.
    - column_range (str, optional): Excel column range (e.g., 'A:O'). Defaults to None.
    - start_row (int): The 1-indexed row number where the data starts. Defaults to 1.
    - end_row (int): Maximum number of rows to read. Defaults to 25000.

    Returns:
    - pd.DataFrame: DataFrame containing the table data.
    """
    try:
        df = pd.read_excel(
            excel_file_path, 
            sheet_name=sheet_name, 
            header=0,
            usecols=column_range, 
            nrows=end_row, 
            skiprows=range(1, start_row)
        )
        
        # Consistent handling of NaNs
        df = df.replace(np.nan, None)
        logger.info(f"Successfully read table from {excel_file_path} [{sheet_name}] with {len(df)} rows.")
        return df

    except Exception as e:
        msg = f"Error reading Excel file: {str(e)}. Check parameters: column_range='{column_range}', start_row={start_row}."
        logger.error(msg)
        raise ValueError(msg)

def update_excel_from_dataframe(
    excel_file_path: str, 
    sheet_name: str, 
    columns_range: str,
    df: pd.DataFrame, 
    df_to_excel_cols_names_dict: dict[str, str] | None = None
) -> None:
    """
    Update specific columns in an existing Excel sheet based on DataFrame content.

    This function clears the old data in the specified range (from row 2 down) 
    before writing the new DataFrame values.

    Parameters:
    - excel_file_path (str): Path to the Excel file.
    - sheet_name (str): Sheet to update.
    - columns_range (str): Range of columns to target (e.g., 'A:O').
    - df (pd.DataFrame): Data source for the update.
    - df_to_excel_cols_names_dict (Dict[str, str], optional): Mapping to rename DF columns to match Excel headers.

    Returns:
    - None
    """
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(excel_file_path)

        if sheet_name not in wb.sheetnames:
            logger.error(f"Sheet '{sheet_name}' not found in {excel_file_path}.")
            return

        sheet = wb[sheet_name]

        # Map column names to letters
        cols_name_dict = {col[0].value.strip(): col[0].column_letter for col in sheet[columns_range]}

        # Clear existing data from row 2 upwards in the specified columns
        for row_idx in range(2, sheet.max_row + 1):
            for col_cell in sheet[columns_range]:
                col_let = col_cell[0].column_letter
                sheet[f"{col_let}{row_idx}"].value = None

        # Apply column renaming if requested
        if df_to_excel_cols_names_dict:
            df = df.rename(columns=df_to_excel_cols_names_dict)

        # Write data row by row
        for col_name in df.columns:
            col_letter = cols_name_dict.get(col_name.strip())
            if not col_letter:
                continue

            for idx, val in enumerate(df[col_name]):
                excel_row = idx + 2
                sheet[f"{col_letter}{excel_row}"] = val

        wb.save(excel_file_path)
        logger.info(f"Updated Excel file '{excel_file_path}' sheet '{sheet_name}' successfully.")

    except Exception as e:
        logger.error(f"Failed to update excel from dataframe: {str(e)}")

def replace_sheet_content(excel_file_path: str, sheet_name: str, df: pd.DataFrame) -> None:
    """
    Replace the entire content of a specific Excel sheet with data from a DataFrame.

    If the sheet exists, it is deleted and recreated. If the file does not exist, a new one is created.

    Parameters:
    - excel_file_path (str): Path to the Excel workbook.
    - sheet_name (str): Name of the sheet to populate.
    - df (pd.DataFrame): DataFrame containing the new content.

    Returns:
    - None
    """
    try:
        try:
            wb = load_workbook(excel_file_path)
        except FileNotFoundError:
            logger.info(f"Creating new workbook at {excel_file_path}")
            wb = Workbook()
            # Remove default 'Sheet' if creating new
            if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
                wb.remove(wb['Sheet'])

        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])

        ws = wb.create_sheet(title=sheet_name)

        # Header
        for col_num, header in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_num, value=header)

        # Content
        for row_idx, row_vals in enumerate(df.values, 2):
            for col_idx, val in enumerate(row_vals, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        wb.save(excel_file_path)
        logger.info(f"Successfully replaced content in sheet '{sheet_name}' of '{excel_file_path}'.")

    except Exception as e:
        logger.error(f"Failed to replace sheet content: {str(e)}")
